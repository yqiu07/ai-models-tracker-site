"""
搜狐号文章列表爬虫（时间段模式）
================================
使用 Selenium 模拟浏览器滚动加载，按时间段提取文章卡片信息，
自动抓取全文，并导出为 JSON/CSV/Excel。

用法:
    python crawl_sohu.py --since 20260417 --until 20260423
    python crawl_sohu.py --since 20260417                    # until 默认为今天
    python crawl_sohu.py                                     # 默认最近 7 天

输出:
    腾讯研究院文章列表.csv / .xlsx（文章列表）
    articles_fulltext.json（带全文的 JSON，供后续 LLM 提取模型信息）

依赖:
    pip install selenium pandas openpyxl requests beautifulsoup4
"""

import argparse
import csv
import json
import os
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# ============================================================
#  配置区 - 修改以下变量即可适配不同搜狐号
# ============================================================

# 目标搜狐号主页 URL
TARGET_URL = (
    "https://mp.sohu.com/profile?xpt="
    "bGl1amluc29uZzIwMDBAMTI2LmNvbQ=="
    "&spm=smpc.content.author.1.1776311557025E5eBJ3T"
)

# 输出目录
OUTPUT_DIR = Path(r"D:\yuwang\action\TXresearch")

# 滚动控制：连续多少次无新文章后放弃
MAX_SCROLL_RETRIES = 15

# 每次滚动后等待的秒数（太短可能加载不出来）
SCROLL_PAUSE_SECONDS = 2.5

# 文章卡片的 CSS 选择器（搜狐号页面的两种卡片类型）
CARD_SELECTORS = "div.TPLTextFeedItem, div.TPLImageTextFeedItem"
CARD_SELECTORS_FALLBACK = (
    "div.TPLPicFeedItem, div.TPLVideoFeedItem, "
    "div.TPLTextFeedItem, div.TPLImageTextFeedItem"
)

# 全文抓取配置
FULLTEXT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
}
FULLTEXT_DELAY_SECONDS = 1.0  # 每篇文章抓取间隔
# ── 时间解析 ─────────────────────────────────────────────────

# 搜狐号卡片上的相对时间模式
_RELATIVE_TIME_PATTERNS = [
    (re.compile(r"(\d+)\s*分钟前"), lambda m: timedelta(minutes=int(m.group(1)))),
    (re.compile(r"(\d+)\s*小时前"), lambda m: timedelta(hours=int(m.group(1)))),
    (re.compile(r"昨天"), lambda _: timedelta(days=1)),
    (re.compile(r"前天"), lambda _: timedelta(days=2)),
    (re.compile(r"(\d+)\s*天前"), lambda m: timedelta(days=int(m.group(1)))),
]

# 标题中的日期模式（如"AI速递 20260320"）
_TITLE_DATE_PATTERN = re.compile(r"(\d{8})")


def parse_article_date(title: str, card_time: str) -> datetime | None:
    """从文章标题或卡片时间解析出日期。

    优先级：
      1. 标题中的 8 位日期（如"AI速递 20260320"中的 20260320）
      2. 卡片上的绝对日期（如"2026-04-20"、"04-20"）
      3. 卡片上的相对时间（如"3小时前"、"昨天"）
    """
    now = datetime.now()

    # 1. 标题中的 YYYYMMDD
    title_match = _TITLE_DATE_PATTERN.search(title)
    if title_match:
        try:
            return datetime.strptime(title_match.group(1), "%Y%m%d")
        except ValueError:
            pass

    if not card_time:
        return None

    # 2. 卡片上的绝对日期格式
    for date_format in ("%Y-%m-%d", "%m-%d", "%Y/%m/%d", "%m月%d日"):
        try:
            parsed = datetime.strptime(card_time.strip(), date_format)
            if parsed.year == 1900:  # 缺少年份
                parsed = parsed.replace(year=now.year)
            return parsed
        except ValueError:
            continue

    # 3. 卡片上的相对时间
    for pattern, delta_func in _RELATIVE_TIME_PATTERNS:
        match = pattern.search(card_time)
        if match:
            return now - delta_func(match)

    return None


def date_to_int(date: datetime) -> int:
    """将 datetime 转为 YYYYMMDD 整数形式（用于与 --since/--until 比较）。"""
    return int(date.strftime("%Y%m%d"))


# ── 页面解析 ─────────────────────────────────────────────────

def extract_articles(driver):
    """从当前页面提取所有文章卡片信息，返回去重后的文章列表。"""
    article_elements = driver.find_elements(By.CSS_SELECTOR, CARD_SELECTORS)

    if not article_elements:
        article_elements = driver.find_elements(By.CSS_SELECTOR, CARD_SELECTORS_FALLBACK)

    if article_elements:
        print(f"[Info] 找到 {len(article_elements)} 个文章卡片")

    articles = []
    seen_titles = set()
    for element in article_elements:
        parsed = _parse_single_card(element)
        if parsed and parsed["title"] not in seen_titles:
            seen_titles.add(parsed["title"])
            articles.append(parsed)

    return articles


def scroll_and_collect(driver, since_int: int, until_int: int):
    """滚动页面并按时间段收集文章。

    Args:
        since_int: 起始日期（YYYYMMDD 整数，如 20260417）
        until_int: 截止日期（YYYYMMDD 整数，如 20260423）

    Returns:
        在时间窗口内的文章列表
    """
    all_articles = []
    matched_articles = []
    seen_titles = set()
    no_new_count = 0
    reached_before_since = False

    print(f"[Start] 开始爬取，时间窗口: {since_int} ~ {until_int}")

    while no_new_count < MAX_SCROLL_RETRIES and not reached_before_since:
        current_articles = extract_articles(driver)
        new_count = 0

        for article in current_articles:
            if article["title"] in seen_titles:
                continue
            seen_titles.add(article["title"])
            all_articles.append(article)
            new_count += 1

            # 解析文章日期
            article_date = parse_article_date(article["title"], article["publish_time"])
            if article_date:
                article_date_int = date_to_int(article_date)
                article["parsed_date"] = article_date.strftime("%Y-%m-%d")
                article["date_int"] = article_date_int

                if article_date_int < since_int:
                    # 文章日期早于起始时间 → 停止爬取
                    print(f"[Stop] 文章 '{article['title'][:40]}' "
                          f"日期={article_date_int} < since={since_int}，停止")
                    reached_before_since = True
                    break
                elif article_date_int > until_int:
                    # 文章日期晚于截止时间 → 跳过，继续滚动
                    print(f"[Skip] 文章 '{article['title'][:40]}' "
                          f"日期={article_date_int} > until={until_int}，跳过")
                else:
                    # 在时间窗口内 → 收集
                    matched_articles.append(article)
                    print(f"[Match] '{article['title'][:40]}' "
                          f"日期={article_date_int} ✅")
            else:
                # 无法解析日期 → 也收集（保守策略，避免遗漏）
                matched_articles.append(article)
                print(f"[Match?] '{article['title'][:40]}' 日期未知，保守收集")

        if new_count > 0:
            print(f"[Progress] 新增 {new_count} 篇，"
                  f"匹配 {len(matched_articles)} 篇，"
                  f"累计扫描 {len(all_articles)} 篇")
            no_new_count = 0
        else:
            no_new_count += 1
            print(f"[Wait] 未发现新文章 ({no_new_count}/{MAX_SCROLL_RETRIES})")

        if reached_before_since:
            break

        # 滚动到页面底部
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(SCROLL_PAUSE_SECONDS)

        # 尝试点击"加载更多"按钮
        for load_sel in [
            "button.load-more", "a.load-more", "div.load-more",
            "[class*='load-more']", "[class*='loadmore']",
        ]:
            try:
                load_btn = driver.find_element(By.CSS_SELECTOR, load_sel)
                if load_btn.is_displayed():
                    load_btn.click()
                    print("[Info] 点击了'加载更多'按钮")
                    time.sleep(SCROLL_PAUSE_SECONDS)
                    break
            except Exception:
                continue

    print(f"[Done] 滚动结束，扫描 {len(all_articles)} 篇，"
          f"匹配时间窗口 {len(matched_articles)} 篇")
    return matched_articles


# ── 全文抓取 ─────────────────────────────────────────────────

def fetch_fulltext(url: str) -> str:
    """抓取单篇搜狐文章的正文内容（requests + BeautifulSoup 三级降级）。"""
    try:
        response = requests.get(url, headers=FULLTEXT_HEADERS, timeout=30)
        response.encoding = "utf-8"
        soup = BeautifulSoup(response.text, "html.parser")

        article_tag = soup.find("article")
        if article_tag:
            paragraphs = article_tag.find_all("p")
        else:
            content_div = soup.find("div", class_=re.compile(r"article|content", re.I))
            if content_div:
                paragraphs = content_div.find_all("p")
            else:
                paragraphs = soup.find_all("p")

        text_parts = [p.get_text(strip=True) for p in paragraphs if len(p.get_text(strip=True)) > 5]
        full_text = "\n".join(text_parts)
        return full_text if len(full_text) > 50 else f"[内容过短] {full_text}"
    except Exception as exc:
        return f"[爬取失败] {exc}"


def fetch_all_fulltexts(articles: list[dict]) -> list[dict]:
    """批量抓取文章全文，返回带全文字段的文章列表。"""
    total = len(articles)
    for idx, article in enumerate(articles, 1):
        url = article.get("link", "")
        if not url:
            article["fulltext"] = "[无链接]"
            continue
        print(f"  [{idx}/{total}] 抓取全文: {article['title'][:40]}...")
        article["fulltext"] = fetch_fulltext(url)
        content_len = len(article["fulltext"])
        status = "✓" if content_len > 100 else "⚠"
        print(f"  {status} {content_len} 字")
        if idx < total:
            time.sleep(FULLTEXT_DELAY_SECONDS)
    return articles


# ── 保存结果 ─────────────────────────────────────────────────

def save_results(articles, output_dir, since_int, until_int):
    """保存结果为 JSON（带全文）+ CSV + Excel（带格式美化）。"""
    _ensure_dir(output_dir)

    tag = f"{since_int}-{until_int}"
    json_path = output_dir / f"articles_{tag}.json"
    csv_path = output_dir / "腾讯研究院文章列表.csv"
    xlsx_path = output_dir / "腾讯研究院文章列表.xlsx"

    # 添加序号
    for index, article in enumerate(articles, 1):
        article["序号"] = index

    # 保存 JSON（带全文，供后续 LLM 提取模型信息）
    json_data = []
    for article in articles:
        json_data.append({
            "序号": article["序号"],
            "标题": article["title"],
            "简介": article.get("summary", ""),
            "链接": article.get("link", ""),
            "发布时间": article.get("parsed_date", article.get("publish_time", "")),
            "全文": article.get("fulltext", ""),
        })
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    print(f"[Save] JSON（含全文）已保存: {json_path}")

    # 保存 CSV
    fieldnames = ["序号", "title", "summary", "link", "publish_time", "read_count", "comment_count"]
    header_names = ["序号", "标题", "简介", "链接", "发布时间", "阅读数", "评论数"]
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header_names)
        for article in articles:
            writer.writerow([article.get(field, "") for field in fieldnames])
    print(f"[Save] CSV 已保存: {csv_path}")

    # 保存 Excel（带格式美化）
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        data_for_df = []
        for article in articles:
            data_for_df.append({
                "序号": article["序号"],
                "标题": article["title"],
                "简介": article.get("summary", ""),
                "链接": article.get("link", ""),
                "发布时间": article.get("parsed_date", article.get("publish_time", "")),
                "阅读数": article.get("read_count", ""),
                "评论数": article.get("comment_count", ""),
            })
        df = pd.DataFrame(data_for_df)
        df.to_excel(xlsx_path, index=False, engine="openpyxl")

        # 美化 Excel
        workbook = load_workbook(xlsx_path)
        worksheet = workbook.active

        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(vertical="center", wrap_text=True)
        link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        column_widths = {"A": 6, "B": 40, "C": 60, "D": 50, "E": 14, "F": 10, "G": 10}
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx in range(2, worksheet.max_row + 1):
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                if col_idx == 4 and cell.value:
                    cell.font = link_font
                    cell.hyperlink = str(cell.value)
            if row_idx % 2 == 0:
                for col_idx in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )

        worksheet.freeze_panes = "A2"
        workbook.save(xlsx_path)
        print(f"[Save] Excel 已保存（含格式美化）: {xlsx_path}")
    except ImportError:
        print("[Warn] pandas/openpyxl 未安装，跳过 Excel 导出")

    return json_path


# ── CLI 入口 ─────────────────────────────────────────────────

def parse_args():
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(
        description="搜狐号文章爬虫（按时间段模式）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "示例:\n"
            "  python crawl_sohu.py --since 20260417 --until 20260423\n"
            "  python crawl_sohu.py --since 20260417\n"
            "  python crawl_sohu.py                  # 默认最近7天\n"
            "  python crawl_sohu.py --no-fulltext     # 只爬列表，不抓全文\n"
        ),
    )
    parser.add_argument(
        "--since", type=str, default=None,
        help="起始日期（YYYYMMDD 格式，如 20260417）",
    )
    parser.add_argument(
        "--until", type=str, default=None,
        help="截止日期（YYYYMMDD 格式，如 20260423），默认今天",
    )
    parser.add_argument(
        "--no-fulltext", action="store_true",
        help="跳过全文抓取（仅爬取文章列表）",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    # 计算时间窗口
    today = datetime.now()
    if args.until:
        until_int = int(args.until)
    else:
        until_int = date_to_int(today)
    if args.since:
        since_int = int(args.since)
    else:
        since_int = date_to_int(today - timedelta(days=7))

    print("=" * 60)
    print("  搜狐号文章爬虫（时间段模式）")
    print(f"  时间窗口: {since_int} ~ {until_int}")
    print("=" * 60)

    driver = create_driver()
    try:
        print(f"\n[Open] 正在打开页面: {TARGET_URL}")
        driver.get(TARGET_URL)
        time.sleep(5)
        print(f"[Info] 页面标题: {driver.title}")

        _ensure_dir(OUTPUT_DIR)

        # 阶段 1: 按时间段滚动收集文章列表
        print(f"\n{'='*40}")
        print("  阶段 1: 爬取文章列表")
        print(f"{'='*40}")
        articles = scroll_and_collect(driver, since_int, until_int)

    finally:
        driver.quit()
        print("[Done] 浏览器已关闭")

    if not articles:
        print("\n[Error] 未能爬取到任何文章")
        print("[Hint] 可能原因：时间窗口内无文章、页面结构变化、网络问题")
        return

    # 阶段 2: 抓取全文
    if not args.no_fulltext:
        print(f"\n{'='*40}")
        print(f"  阶段 2: 抓取 {len(articles)} 篇文章全文")
        print(f"{'='*40}")
        articles = fetch_all_fulltexts(articles)

    # 阶段 3: 保存结果
    print(f"\n{'='*40}")
    print("  阶段 3: 保存结果")
    print(f"{'='*40}")
    json_path = save_results(articles, OUTPUT_DIR, since_int, until_int)

    # 打印结果摘要
    print(f"\n{'='*60}")
    print(f"  📊 爬取完成！")
    print(f"  时间窗口: {since_int} ~ {until_int}")
    print(f"  匹配文章: {len(articles)} 篇")
    if not args.no_fulltext:
        ok_count = sum(
            1 for a in articles
            if a.get("fulltext", "").startswith("[") is False and len(a.get("fulltext", "")) > 100
        )
        print(f"  全文抓取: {ok_count}/{len(articles)} 成功")
    print(f"  JSON 输出: {json_path}")
    print(f"{'='*60}")

    # 列出匹配的文章
    print(f"\n  匹配文章列表:")
    for article in articles:
        date_str = article.get("parsed_date", "未知")
        print(f"    {date_str} | {article['title'][:50]}")


if __name__ == "__main__":
    main()